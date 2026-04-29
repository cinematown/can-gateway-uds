#!/usr/bin/env python3
"""Generate a DBC file from a VW KMatrix Excel workbook."""

from __future__ import annotations

import argparse
import re
import unicodedata
from collections import OrderedDict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


NODE_COLUMNS = range(34, 79)


DBC_HEADER = """VERSION "Generated from PQ35_46 ACAN KMatrix"


NS_ :
\tNS_DESC_
\tCM_
\tBA_DEF_
\tBA_
\tVAL_
\tCAT_DEF_
\tCAT_
\tFILTER
\tBA_DEF_DEF_
\tEV_DATA_
\tENVVAR_DATA_
\tSGTYPE_
\tSGTYPE_VAL_
\tBA_DEF_SGTYPE_
\tBA_SGTYPE_
\tSIG_TYPE_REF_
\tVAL_TABLE_
\tSIG_GROUP_
\tSIG_VALTYPE_
\tSIGTYPE_VALTYPE_
\tSG_MUL_VAL_

BS_:

"""


@dataclass
class Signal:
    name: str
    start_bit: int
    length: int
    scale: Decimal
    offset: Decimal
    minimum: Decimal
    maximum: Decimal
    unit: str
    receivers: list[str]
    values: list[tuple[int, str]] = field(default_factory=list)
    comment: str = ""
    send_type: str = ""


@dataclass
class Message:
    frame_id: int
    name: str
    dlc: int
    cycle_time: int | None = None
    send_type: str = ""
    signals: list[Signal] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="KMatrix workbook path")
    parser.add_argument("dbc", type=Path, help="DBC output path")
    parser.add_argument(
        "--sheet",
        default="PQ35_46_ACAN ",
        help="Workbook sheet containing messages and signals",
    )
    return parser.parse_args()


def to_ascii(value: object) -> str:
    text = "" if value is None else str(value)
    replacements = {
        "Ä": "Ae",
        "Ö": "Oe",
        "Ü": "Ue",
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
        "°": "deg",
        "µ": "u",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def dbc_name(value: object, fallback: str) -> str:
    text = to_ascii(value).strip()
    text = re.sub(r"\W+", "_", text)
    text = text.strip("_")
    if not text:
        text = fallback
    if text[0].isdigit():
        text = "_" + text
    return text


def dbc_string(value: object) -> str:
    text = to_ascii(value)
    text = re.sub(r"\s+", " ", text.replace("\r", "\n").replace("\n", " | ")).strip()
    return text.replace("\\", "\\\\").replace('"', r"\"")


def decimal_or_default(value: object, default: Decimal) -> Decimal:
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, AttributeError):
        return default


def int_or_none(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(Decimal(str(value).replace(",", ".").strip()))
    except (InvalidOperation, ValueError):
        return None


def format_decimal(value: Decimal) -> str:
    if value == 0:
        return "0"
    if value == value.to_integral_value():
        return str(int(value))
    return format(value, "f").rstrip("0").rstrip(".")


def split_cell_lines(value: object) -> list[str]:
    if value is None or value == "":
        return []
    return str(value).replace("\r\n", "\n").replace("\r", "\n").split("\n")


def value_table(raw_values: object, descriptions: object) -> list[tuple[int, str]]:
    values = [line.strip() for line in split_cell_lines(raw_values)]
    descs = split_cell_lines(descriptions)
    pairs: list[tuple[int, str]] = []
    for index, raw_value in enumerate(values):
        if not raw_value:
            continue
        parsed = int_or_none(raw_value)
        if parsed is None:
            continue
        desc = descs[index] if index < len(descs) else ""
        pairs.append((parsed, dbc_string(desc)))
    return pairs


def physical_limits(row: tuple[object, ...], length: int, scale: Decimal, offset: Decimal) -> tuple[Decimal, Decimal]:
    raw_min = decimal_or_default(row[24] if len(row) > 24 else None, Decimal("NaN"))
    raw_max = decimal_or_default(row[25] if len(row) > 25 else None, Decimal("NaN"))

    if raw_min.is_nan() or raw_max.is_nan():
        logical_values = [int_or_none(line.strip()) for line in split_cell_lines(row[30] if len(row) > 30 else None)]
        numeric_values = [Decimal(value) for value in logical_values if value is not None]
        if numeric_values:
            raw_min = min(numeric_values)
            raw_max = max(numeric_values)
        elif length <= 32:
            raw_min = Decimal(0)
            raw_max = Decimal((1 << length) - 1)
        else:
            raw_min = Decimal(0)
            raw_max = Decimal(0)

    phys = [raw_min * scale + offset, raw_max * scale + offset]
    return min(phys), max(phys)


def receivers_from_row(row: tuple[object, ...], node_names: dict[int, str]) -> list[str]:
    receivers: list[str] = []
    for column in NODE_COLUMNS:
        value = row[column - 1] if len(row) >= column else None
        if isinstance(value, str) and value.startswith("E"):
            receivers.append(node_names[column])
    return receivers or ["XXX"]


def load_messages(xlsx_path: Path, sheet_name: str) -> tuple[OrderedDict[int, Message], list[str]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]

    header_row = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
    node_names = {
        column: dbc_name(header_row[column - 1], f"Node_{column}") for column in NODE_COLUMNS
    }

    messages: OrderedDict[int, Message] = OrderedDict()
    signal_names_by_message: dict[int, set[str]] = {}

    for row_index, row in enumerate(worksheet.iter_rows(min_row=5, values_only=True), start=5):
        message_name_raw = row[0] if len(row) > 0 else None
        frame_id = int_or_none(row[2] if len(row) > 2 else None)
        signal_name_raw = row[12] if len(row) > 12 else None
        start_byte = int_or_none(row[13] if len(row) > 13 else None)
        bit_in_byte = int_or_none(row[14] if len(row) > 14 else None)
        signal_length = int_or_none(row[15] if len(row) > 15 else None)

        if message_name_raw in (None, "") or frame_id is None:
            continue

        message_name = dbc_name(message_name_raw, f"Message_{frame_id}")
        dlc = int_or_none(row[4] if len(row) > 4 else None) or 8
        message = messages.get(frame_id)
        if message is None:
            message = Message(
                frame_id=frame_id,
                name=message_name,
                dlc=dlc,
                cycle_time=int_or_none(row[5] if len(row) > 5 else None),
                send_type=dbc_string(row[10] if len(row) > 10 else ""),
            )
            messages[frame_id] = message
            signal_names_by_message[frame_id] = set()

        if signal_name_raw in (None, "", "void") or start_byte is None or bit_in_byte is None or signal_length is None:
            continue

        signal_name = dbc_name(signal_name_raw, f"Signal_{row_index}")
        if signal_name in signal_names_by_message[frame_id]:
            signal_name = f"{signal_name}_{row_index}"
        signal_names_by_message[frame_id].add(signal_name)

        start_bit = (start_byte - 1) * 8 + bit_in_byte
        scale = decimal_or_default(row[29] if len(row) > 29 else None, Decimal(1))
        offset = decimal_or_default(row[28] if len(row) > 28 else None, Decimal(0))
        minimum, maximum = physical_limits(row, signal_length, scale, offset)

        message.signals.append(
            Signal(
                name=signal_name,
                start_bit=start_bit,
                length=signal_length,
                scale=scale,
                offset=offset,
                minimum=minimum,
                maximum=maximum,
                unit=dbc_string(row[27] if len(row) > 27 else ""),
                receivers=receivers_from_row(row, node_names),
                values=value_table(row[30] if len(row) > 30 else None, row[31] if len(row) > 31 else None),
                comment=dbc_string(row[79] if len(row) > 79 else ""),
                send_type=dbc_string(row[16] if len(row) > 16 else ""),
            )
        )

    nodes = ["XXX", *sorted(set(node_names.values()))]
    return messages, nodes


def write_dbc(messages: OrderedDict[int, Message], nodes: Iterable[str], output_path: Path) -> None:
    lines: list[str] = [DBC_HEADER]
    lines.append("BU_: " + " ".join(nodes))
    lines.append("")
    lines.append("")

    for message in messages.values():
        lines.append(f"BO_ {message.frame_id} {message.name}: {message.dlc} XXX")
        for signal in message.signals:
            receivers = ",".join(signal.receivers)
            lines.append(
                " SG_ {name} : {start}|{length}@1+ ({scale},{offset}) [{minimum}|{maximum}] \"{unit}\" {receivers}".format(
                    name=signal.name,
                    start=signal.start_bit,
                    length=signal.length,
                    scale=format_decimal(signal.scale),
                    offset=format_decimal(signal.offset),
                    minimum=format_decimal(signal.minimum),
                    maximum=format_decimal(signal.maximum),
                    unit=signal.unit,
                    receivers=receivers,
                )
            )
        lines.append("")

    lines.append('CM_ "Generated from VW PQ35_46 ACAN KMatrix V5.20.6F, 2016-05-30.";')
    for message in messages.values():
        for signal in message.signals:
            if signal.comment:
                lines.append(f'CM_ SG_ {message.frame_id} {signal.name} "{signal.comment}";')

    lines.append('BA_DEF_ BO_ "GenMsgCycleTime" INT 0 65535;')
    lines.append('BA_DEF_ BO_ "GenMsgSendType" STRING ;')
    lines.append('BA_DEF_ SG_ "GenSigSendType" STRING ;')
    lines.append('BA_DEF_DEF_ "GenMsgCycleTime" 0;')
    lines.append('BA_DEF_DEF_ "GenMsgSendType" "";')
    lines.append('BA_DEF_DEF_ "GenSigSendType" "";')

    for message in messages.values():
        if message.cycle_time is not None:
            lines.append(f'BA_ "GenMsgCycleTime" BO_ {message.frame_id} {message.cycle_time};')
        if message.send_type:
            lines.append(f'BA_ "GenMsgSendType" BO_ {message.frame_id} "{message.send_type}";')
        for signal in message.signals:
            if signal.send_type:
                lines.append(f'BA_ "GenSigSendType" SG_ {message.frame_id} {signal.name} "{signal.send_type}";')

    for message in messages.values():
        for signal in message.signals:
            if signal.values:
                values = " ".join(f'{value} "{description}"' for value, description in reversed(signal.values))
                lines.append(f"VAL_ {message.frame_id} {signal.name} {values} ;")

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    args = parse_args()
    messages, nodes = load_messages(args.xlsx, args.sheet)
    write_dbc(messages, nodes, args.dbc)
    print(f"Wrote {args.dbc} ({len(messages)} messages)")


if __name__ == "__main__":
    main()
